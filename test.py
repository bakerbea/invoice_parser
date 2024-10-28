import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import shutil
from zipfile import ZipFile

# Define output directory for saving generated files
OUTPUT_DIR = 'orders_output'
ZIP_FILE = 'orders_output.zip'

# Clear out the existing `orders_output` folder before generating new files
if os.path.exists(OUTPUT_DIR):
    shutil.rmtree(OUTPUT_DIR)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Streamlit interface title
st.title("Order Processing App")

# Dropdown for selecting file type
file_type = st.selectbox(
    "Select the type of file you are uploading:",
    ["Lazada", "Zalora", "Shopee", "TikTok", "JDO"]
)

# File uploader widget
uploaded_file = st.file_uploader("Upload an Excel file", type="xlsx")

if uploaded_file is not None:
    # Load the uploaded Excel file
    df = pd.read_excel(uploaded_file)
    columns = {}
    
    # Case-based functionality based on the selected file type
    if file_type == "Lazada":
        columns = {
            'customer_name': 'customerName',
            'province': 'shippingCountry',
            'styleclrsize': 'sellerSku',
            'unit_price': 'paidPrice',
            'qty': None,
            'order_date': 'createTime',
            'order_id': 'orderNumber',
            'style_name': 'itemName'
        }
    elif file_type == "Shopee":
        columns = {
            'customer_name': 'Name',
            'province': 'City',
            'styleclrsize': 'SKU Reference No.',
            'unit_price': 'Deal Price',
            'qty': 'Quantity',
            'order_date': 'Order Creation Date',
            'order_id': 'Order ID',
            'style_name': 'Product Name'
        }
    elif file_type == "Zalora":
        columns = {
            'customer_name': 'Customer Name',
            'province': 'Shipping City',
            'styleclrsize': 'Seller SKU',
            'unit_price': 'Paid Price',
            'qty': None,
            'order_date': 'Created at',
            'order_id': 'Order Number',
            'style_name': 'Item Name'
        }
    elif file_type == "TikTok":
        columns = {
            'customer_name': 'Buyer Username',
            'province': 'Province',
            'styleclrsize': 'Seller SKU',
            'unit_price': 'SKU Subtotal After Discount',
            'qty': 'Quantity',
            'order_date': 'Created Time',
            'order_id': 'Order ID',
            'style_name': 'Product Name'
        }
    elif file_type == "JDO":
        columns = {
            'customer_name': 'CUSTOMER NAME',
            'province': 'PROVINCE',
            'styleclrsize': 'STYLECLRSIZE',
            'unit_price': 'UNIT PRICE',
            'qty': 'QTY',
            'order_date': 'ORDER DATE',
            'order_id': None,
            'style_name': None
        }

    # Extract relevant columns
    df_filtered = df[[value for value in columns.values() if value is not None]]
    grouped_orders = df_filtered.groupby([columns['customer_name'], columns['order_date']])

    def sanitize_filename(filename):
        return filename.replace(':', '-').replace('/', '-').replace('\\', '-')

    def write_order_to_template(customer_name, province, order_id, order_data, output_file, start_item=0, end_item=None):
        wb = Workbook()
        ws = wb.active

        ws.page_setup.orientation = 'portrait'
        ws.page_setup.scale = 99
        ws.print_area = 'B1:K40'
        ws.page_margins.top = 0.0
        ws.page_margins.bottom = 0.74
        ws.page_margins.left = 0.0787
        ws.page_margins.right = 0.118
        ws.page_margins.header = 0.315
        ws.page_margins.footer = 0.315

        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=10)

        # Define column widths 
        column_widths = {
            'A': 1.854,     
            'B': 4.326,     
            'C': 5.253,     
            'D': 11.433,    
            'E': 4.635,     
            'F': 11.2785,   
            'G': 7.416,     
            'H': 16.995,    
            'I': 7.416,     
            'J': 11.433,    
            'K': 13.978 
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        row_heights = {1: 11.45, 2: 11.45, 3: 11.45, 4: 11.45, 5: 14.25, 6: 11.45,
                       7: 11.45, 8: 13.5, 9: 15.0, 10: 4.5, 11: 12.0, 12: 17.25,
                       13: 21.75, 16: 12.75, 17: 11.25, 18: 11.25, 19: 11.25,
                       20: 11.25, 21: 11.25, 22: 11.25, 23: 11.25, 24: 11.25,
                       25: 11.25, 26: 11.25, 27: 9.0, 28: 11.5, 29: 9.5, 30: 10.5,
                       31: 10.5, 32: 11.5, 33: 11.5, 34: 8.5, 35: 9.75, 36: 7.5,
                       37: 12.0, 38: 15.0, 39: 18.75, 40: 30.0}

        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        ws['D9'] = customer_name
        ws['D11'] = province
        ws['K11'] = datetime.now().strftime("%B %d, %Y")
        ws.merge_cells('I13:J13')
        ws['I13'] = str(order_id)
        ws.merge_cells('E14:F14')
        ws['E14'] = file_type

        start_row = 17
        total_value = 0
        p_count = 0

        if end_item is None:
            end_item = len(order_data)

        for index, row in order_data.iloc[start_item:end_item].iterrows():
            qty = int(row[columns['qty']]) if columns['qty'] is not None else 1
            style_total = qty * float(row[columns['unit_price']]) if file_type != 'TikTok' \
                else qty * float(row[columns['unit_price']].replace("PHP", "").replace(',', '').strip())
            total_value += style_total
            p_count += qty

            ws.merge_cells(f'C{start_row}:D{start_row}')
            ws[f'C{start_row}'] = row[columns['styleclrsize']]
            ws[f'F{start_row}'] = row[columns['style_name']] if columns['style_name'] is not None else ''
            ws[f'I{start_row}'] = qty
            ws[f'J{start_row}'] = row[columns['unit_price']]
            ws[f'K{start_row}'] = str(style_total)

            start_row += 1

        num_styles = end_item - start_item
        num_dashes = max(10 - num_styles, 0)

        for i in range(num_dashes):
            ws[f'J{start_row}'] = '-'
            start_row += 1

        start_row += 1
        ws[f'H{start_row}'] = round(total_value / 1.12, 2)
        start_row += 4
        ws[f'H{start_row}'] = total_value
        start_row += 1
        ws[f'H{start_row}'] = total_value - round(total_value / 1.12, 2)
        start_row += 4
        ws[f'I{start_row}'] = p_count
        ws[f'J{start_row}'] = 'P'
        ws[f'K{start_row}'] = total_value

        wb.save(output_file)

    for (customer, date), group in grouped_orders:
        customer_name = group[columns['customer_name']].iloc[0]
        province = group[columns['province']].iloc[0]
        order_id = group[columns['order_id']].iloc[0] if columns['order_id'] is not None else ''
        date_str = str(date).replace(':', '-').replace(' ', '_').replace('/','-')
        customer_str = sanitize_filename(customer_name)

        num_items = len(group)
        batch_size = 10
        batch_number = 1
        for start_item in range(0, num_items, batch_size):
            end_item = min(start_item + batch_size, num_items)
            file_name = f"{customer_str}_{date_str}_batch{batch_number}.xlsx"
            output_file = os.path.join(OUTPUT_DIR, file_name)
            write_order_to_template(customer_name, province, order_id, group, output_file, start_item, end_item)
            batch_number += 1

    with ZipFile(ZIP_FILE, 'w') as zipf:
        for root, dirs, files in os.walk(OUTPUT_DIR):
            for file in files:
                zipf.write(os.path.join(root, file), arcname=file)

    with open(ZIP_FILE, 'rb') as f:
        st.download_button(label="Download All Orders (ZIP)", data=f, file_name="orders_output.zip")
